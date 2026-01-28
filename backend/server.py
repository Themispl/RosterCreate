from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import calendar
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Default color configuration matching the reference image
DEFAULT_SHIFT_COLORS = {
    "7": {"bg": "FF6666", "text": "000000"},      # Morning - Red
    "15": {"bg": "009933", "text": "FFFFFF"},     # Afternoon - Green
    "23": {"bg": "3366CC", "text": "FFFFFF"},     # Night - Blue
    "9": {"bg": "FFFF66", "text": "000000"},      # 9am - Yellow
    "11": {"bg": "6699FF", "text": "000000"},     # 11am - Light Blue
    "12": {"bg": "9966CC", "text": "FFFFFF"},     # 12pm - Purple
    "8": {"bg": "FF9999", "text": "000000"},      # 8am - Pink
    "16": {"bg": "CC0033", "text": "FFFFFF"},     # 16pm - Maroon
    "0": {"bg": "FF9999", "text": "B91C1C"},      # Day Off - Light Pink with red text
    "V": {"bg": "663399", "text": "FFFFFF"},      # Vacation - Dark Purple
    "L": {"bg": "CC6600", "text": "FFFFFF"},      # Leave - Orange
}

# Position sort order
POSITION_ORDER = {"AGSM": 0, "GSC": 1, "GSA": 2, "Welcome Agent": 3}

# Models
class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    last_name: str
    first_name: str
    position: str  # GSC, GSA, AGSM, Welcome Agent
    group: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeCreate(BaseModel):
    last_name: str
    first_name: str
    position: str
    group: Optional[str] = None

class EmployeeUpdate(BaseModel):
    last_name: Optional[str] = None
    first_name: Optional[str] = None
    position: Optional[str] = None
    group: Optional[str] = None

class ColorConfig(BaseModel):
    bg: str
    text: str

class RosterRequest(BaseModel):
    year: int
    month: int
    employees: List[str]
    vacation_days: Dict[str, List[str]] = {}
    leave_days: Dict[str, List[str]] = {}
    custom_colors: Dict[str, ColorConfig] = {}
    view_type: str = "month"  # "month" or "week"
    week_number: Optional[int] = None  # 1-5 for week view

class RosterResponse(BaseModel):
    year: int
    month: int
    roster: Dict[str, Dict[str, str]]
    days_info: List[Dict[str, Any]]
    view_type: str
    week_number: Optional[int] = None

# Employee endpoints
@api_router.post("/employees", response_model=Employee)
async def create_employee(employee: EmployeeCreate):
    emp = Employee(**employee.model_dump())
    doc = emp.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.employees.insert_one(doc)
    return emp

@api_router.get("/employees", response_model=List[Employee])
async def get_employees():
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    for emp in employees:
        if isinstance(emp.get('created_at'), str):
            emp['created_at'] = datetime.fromisoformat(emp['created_at'])
    
    # Sort by position order: AGSM → GSC → GSA → Welcome Agent
    employees.sort(key=lambda x: (POSITION_ORDER.get(x['position'], 99), x['last_name']))
    return employees

@api_router.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(employee_id: str, update: EmployeeUpdate):
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.employees.update_one(
        {"id": employee_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if isinstance(employee.get('created_at'), str):
        employee['created_at'] = datetime.fromisoformat(employee['created_at'])
    return Employee(**employee)

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str):
    result = await db.employees.delete_one({"id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted"}

@api_router.post("/employees/bulk")
async def bulk_create_employees(employees: List[EmployeeCreate]):
    created = []
    for emp_data in employees:
        emp = Employee(**emp_data.model_dump())
        doc = emp.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.employees.insert_one(doc)
        created.append(emp)
    return created

@api_router.post("/employees/import-csv")
async def import_csv(file: UploadFile = File(...)):
    content = await file.read()
    decoded = content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(decoded))
    
    created = []
    for row in reader:
        emp_data = EmployeeCreate(
            last_name=row.get('last_name', row.get('LAST_NAME', row.get('Last Name', ''))),
            first_name=row.get('first_name', row.get('FIRST_NAME', row.get('First Name', ''))),
            position=row.get('position', row.get('POSITION', row.get('Position', 'GSC'))),
            group=row.get('group', row.get('GROUP', row.get('Group', None)))
        )
        emp = Employee(**emp_data.model_dump())
        doc = emp.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.employees.insert_one(doc)
        created.append(emp)
    
    return {"imported": len(created), "employees": created}


def generate_roster(year: int, month: int, employees: List[dict], 
                   vacation_days: Dict[str, List[str]] = {},
                   leave_days: Dict[str, List[str]] = {}) -> Dict[str, Dict[str, str]]:
    """
    Generate a roster following ALL constraints:
    1. 5 work days + 2 days off per week
    2. 11-hour rest rule: No AM→PM transition without day off between
    3. Days off used to transition between shift types
    4. AGSM and Welcome Agent only 9am shifts
    5. Max 5 night shifts per person per month
    6. Days off balanced across staff (not everyone off same day)
    7. Days off per week always consecutive (together) per person
    """
    num_days = calendar.monthrange(year, month)[1]
    roster = {emp['id']: {} for emp in employees}
    
    # Separate employees by position
    fixed_9am = [e for e in employees if e['position'] in ['AGSM', 'Welcome Agent']]
    flexible = [e for e in employees if e['position'] not in ['AGSM', 'Welcome Agent']]
    
    # Assign each employee their off days for each week (staggered for balance)
    # Each employee gets 2 consecutive days off per week, different employees different days
    employee_off_days = {}  # emp_id -> list of (start_weekday) for each week
    
    for i, emp in enumerate(employees):
        emp_id = emp['id']
        # Stagger off days: employee 0 gets Mon-Tue, employee 1 gets Tue-Wed, etc.
        off_start = i % 6  # 0-5 (Mon-Sat as start of 2-day off period)
        employee_off_days[emp_id] = off_start
    
    # Track state per employee
    employee_state = {}
    for emp in employees:
        employee_state[emp['id']] = {
            'last_shift': None,
            'current_shift_type': None,  # 'morning' or 'afternoon'
            'night_shifts_count': 0,
            'in_night_rotation': False,
            'night_days_done': 0,
        }
    
    # Night rotation queue - only flexible employees
    night_queue = list(flexible)
    night_idx = 0
    current_night_worker_id = None
    
    # Process each day
    for day in range(1, num_days + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        date_obj = datetime(year, month, day)
        weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
        week_number = (day - 1) // 7
        
        # First, mark vacation and leave days
        for emp in employees:
            emp_id = emp['id']
            if vacation_days.get(emp_id) and date_str in vacation_days[emp_id]:
                roster[emp_id][date_str] = 'V'
            elif leave_days.get(emp_id) and date_str in leave_days[emp_id]:
                roster[emp_id][date_str] = 'L'
        
        # Process each employee
        for emp in employees:
            emp_id = emp['id']
            
            # Skip if already assigned (vacation/leave)
            if roster[emp_id].get(date_str):
                continue
            
            state = employee_state[emp_id]
            
            # Check if this is an off day for this employee
            off_start = employee_off_days[emp_id]
            is_off_day = weekday == off_start or weekday == (off_start + 1) % 7
            
            # AGSM and Welcome Agent: only 9am shifts
            if emp['position'] in ['AGSM', 'Welcome Agent']:
                if is_off_day:
                    roster[emp_id][date_str] = '0'
                else:
                    roster[emp_id][date_str] = '9'
                    state['last_shift'] = '9'
                continue
            
            # Flexible employees
            # Check if in night rotation
            if state['in_night_rotation']:
                if state['night_days_done'] < 5:
                    # Continue night shift
                    roster[emp_id][date_str] = '23'
                    state['night_days_done'] += 1
                    state['night_shifts_count'] += 1
                    state['last_shift'] = '23'
                    continue
                else:
                    # Finished 5 nights, need 2 days off then switch to afternoon
                    state['in_night_rotation'] = False
                    state['current_shift_type'] = 'afternoon'
                    # Give 2 days off
                    roster[emp_id][date_str] = '0'
                    continue
            
            # Check if this is the designated off day
            if is_off_day:
                roster[emp_id][date_str] = '0'
                continue
            
            # Check if we need to start night rotation for this employee
            if (current_night_worker_id is None and 
                emp_id in [e['id'] for e in night_queue] and
                state['night_shifts_count'] < 5 and
                weekday == 0):  # Start night rotation on Monday
                
                # Check if this employee should do night shift
                if night_queue and night_idx < len(night_queue):
                    if night_queue[night_idx]['id'] == emp_id:
                        current_night_worker_id = emp_id
                        state['in_night_rotation'] = True
                        state['night_days_done'] = 0
                        night_idx += 1
                        
                        roster[emp_id][date_str] = '23'
                        state['night_days_done'] += 1
                        state['night_shifts_count'] += 1
                        state['last_shift'] = '23'
                        continue
            
            # Determine shift type based on week alternation
            if state['current_shift_type'] is None:
                emp_idx = flexible.index(emp) if emp in flexible else 0
                state['current_shift_type'] = 'morning' if (week_number + emp_idx) % 2 == 0 else 'afternoon'
            
            # Check 11-hour rest rule before assigning shift
            last = state['last_shift']
            target_type = state['current_shift_type']
            
            # If switching from morning to afternoon or vice versa, need a day off
            if last == '7' and target_type == 'afternoon':
                # Need transition - use today as off day
                roster[emp_id][date_str] = '0'
                state['current_shift_type'] = 'afternoon'
                continue
            elif last == '15' and target_type == 'morning':
                # Need transition - use today as off day
                roster[emp_id][date_str] = '0'
                state['current_shift_type'] = 'morning'
                continue
            elif last == '23':
                # After night shift, need off or afternoon only
                if target_type == 'morning':
                    roster[emp_id][date_str] = '0'
                    continue
            
            # Assign the shift
            if target_type == 'morning':
                roster[emp_id][date_str] = '7'
                state['last_shift'] = '7'
            else:
                roster[emp_id][date_str] = '15'
                state['last_shift'] = '15'
            
            # Alternate shift type at end of week (Sunday)
            if weekday == 6:
                if state['current_shift_type'] == 'morning':
                    state['current_shift_type'] = 'afternoon'
                else:
                    state['current_shift_type'] = 'morning'
        
        # Reset night worker if they finished
        if current_night_worker_id:
            state = employee_state[current_night_worker_id]
            if not state['in_night_rotation']:
                current_night_worker_id = None
    
    # Second pass: Ensure exactly 2 consecutive off days per week
    for emp in employees:
        emp_id = emp['id']
        emp_roster = roster[emp_id]
        
        # Group by weeks
        for week in range(5):  # Max 5 weeks in a month
            week_start = week * 7 + 1
            week_end = min(week_start + 6, num_days)
            
            if week_start > num_days:
                break
            
            # Count off days in this week
            off_days_in_week = []
            for d in range(week_start, week_end + 1):
                date_str = f"{year}-{month:02d}-{d:02d}"
                if emp_roster.get(date_str) == '0':
                    off_days_in_week.append(d)
            
            # If no off days, add 2 consecutive
            if len(off_days_in_week) == 0:
                off_start = employee_off_days[emp_id]
                for d in range(week_start, week_end + 1):
                    date_obj = datetime(year, month, d)
                    if date_obj.weekday() == off_start:
                        date_str = f"{year}-{month:02d}-{d:02d}"
                        emp_roster[date_str] = '0'
                        if d + 1 <= week_end:
                            next_date = f"{year}-{month:02d}-{d+1:02d}"
                            emp_roster[next_date] = '0'
                        break
            
            # If only 1 off day, add another consecutive
            elif len(off_days_in_week) == 1:
                d = off_days_in_week[0]
                # Try to add next day as off
                if d + 1 <= num_days:
                    next_date = f"{year}-{month:02d}-{d+1:02d}"
                    if emp_roster.get(next_date) not in ['V', 'L']:
                        emp_roster[next_date] = '0'
                # Or previous day
                elif d - 1 >= week_start:
                    prev_date = f"{year}-{month:02d}-{d-1:02d}"
                    if emp_roster.get(prev_date) not in ['V', 'L']:
                        emp_roster[prev_date] = '0'
            
            # If off days not consecutive, make them consecutive
            elif len(off_days_in_week) >= 2:
                # Check if first two are consecutive
                if off_days_in_week[1] - off_days_in_week[0] != 1:
                    # Make the second off day right after the first
                    d = off_days_in_week[0]
                    if d + 1 <= num_days:
                        next_date = f"{year}-{month:02d}-{d+1:02d}"
                        if emp_roster.get(next_date) not in ['V', 'L']:
                            # Remove old second off day
                            old_off = f"{year}-{month:02d}-{off_days_in_week[1]:02d}"
                            if emp_roster.get(old_off) == '0':
                                # Assign a shift to old off day based on current shift type
                                state = employee_state[emp_id]
                                if emp['position'] in ['AGSM', 'Welcome Agent']:
                                    emp_roster[old_off] = '9'
                                elif state['current_shift_type'] == 'morning':
                                    emp_roster[old_off] = '7'
                                else:
                                    emp_roster[old_off] = '15'
                            emp_roster[next_date] = '0'
    
    return roster


@api_router.post("/roster/generate", response_model=RosterResponse)
async def generate_roster_endpoint(request: RosterRequest):
    # Get employees from database
    employees = await db.employees.find(
        {"id": {"$in": request.employees}},
        {"_id": 0}
    ).to_list(1000)
    
    if not employees:
        raise HTTPException(status_code=400, detail="No employees found")
    
    # Sort employees by position order
    employees.sort(key=lambda x: (POSITION_ORDER.get(x['position'], 99), x['last_name']))
    
    roster = generate_roster(
        request.year, 
        request.month, 
        employees,
        request.vacation_days,
        request.leave_days
    )
    
    # Generate days info
    num_days = calendar.monthrange(request.year, request.month)[1]
    weekday_names = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
    days_info = []
    
    # Filter days based on view type
    if request.view_type == "week" and request.week_number:
        start_day = (request.week_number - 1) * 7 + 1
        end_day = min(start_day + 6, num_days)
        day_range = range(start_day, end_day + 1)
    else:
        day_range = range(1, num_days + 1)
    
    for day in day_range:
        date_obj = datetime(request.year, request.month, day)
        days_info.append({
            "day": day,
            "weekday": weekday_names[date_obj.weekday()],
            "date": f"{request.year}-{request.month:02d}-{day:02d}"
        })
    
    return RosterResponse(
        year=request.year,
        month=request.month,
        roster=roster,
        days_info=days_info,
        view_type=request.view_type,
        week_number=request.week_number
    )


@api_router.post("/roster/export-excel")
async def export_excel(request: RosterRequest):
    """Export roster as formatted Excel file"""
    # Get employees
    employees = await db.employees.find(
        {"id": {"$in": request.employees}},
        {"_id": 0}
    ).to_list(1000)
    
    if not employees:
        raise HTTPException(status_code=400, detail="No employees found")
    
    # Sort employees by position order
    employees.sort(key=lambda x: (POSITION_ORDER.get(x['position'], 99), x['last_name']))
    
    # Generate roster
    roster = generate_roster(
        request.year,
        request.month,
        employees,
        request.vacation_days,
        request.leave_days
    )
    
    # Merge custom colors with defaults
    shift_colors = DEFAULT_SHIFT_COLORS.copy()
    for key, color in request.custom_colors.items():
        shift_colors[key] = {"bg": color.bg.replace('#', ''), "text": color.text.replace('#', '')}
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Roster {request.month:02d}-{request.year}"
    
    num_days = calendar.monthrange(request.year, request.month)[1]
    weekday_names = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
    
    # Header styling
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row 1 (day numbers)
    ws.cell(row=1, column=1, value="LAST NAME")
    ws.cell(row=1, column=2, value="FIRST NAME")
    ws.cell(row=1, column=3, value="POSITION")
    
    for day in range(1, num_days + 1):
        col = day + 3
        ws.cell(row=1, column=col, value=day)
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=col).border = thin_border
    
    # Write header row 2 (weekday names)
    for day in range(1, num_days + 1):
        col = day + 3
        date_obj = datetime(request.year, request.month, day)
        ws.cell(row=2, column=col, value=weekday_names[date_obj.weekday()])
        ws.cell(row=2, column=col).font = Font(bold=True, size=8)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col).border = thin_border
    
    # Set column widths - wider for names
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    for day in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(day + 3)].width = 5
    
    # Write employee rows
    row = 3
    current_group = None
    
    for emp in employees:
        # Add group header if group changes
        if emp.get('group') and emp.get('group') != current_group:
            current_group = emp.get('group')
            ws.cell(row=row, column=1, value=current_group)
            ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
            row += 1
        
        ws.cell(row=row, column=1, value=emp['last_name'])
        ws.cell(row=row, column=2, value=emp['first_name'])
        ws.cell(row=row, column=3, value=emp['position'])
        
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
        
        # Write shifts with colors
        emp_roster = roster.get(emp['id'], {})
        for day in range(1, num_days + 1):
            col = day + 3
            date_str = f"{request.year}-{request.month:02d}-{day:02d}"
            shift = emp_roster.get(date_str, '')
            
            cell = ws.cell(row=row, column=col, value=shift)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # Apply color based on shift
            if shift in shift_colors:
                color = shift_colors[shift]
                cell.fill = PatternFill(start_color=color["bg"], end_color=color["bg"], fill_type="solid")
                cell.font = Font(color=color["text"], bold=True)
        
        row += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"roster_{request.year}_{request.month:02d}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/colors")
async def get_colors():
    """Get current color configuration"""
    colors = await db.color_config.find_one({"type": "shift_colors"}, {"_id": 0})
    if colors:
        return colors.get("colors", DEFAULT_SHIFT_COLORS)
    return DEFAULT_SHIFT_COLORS

@api_router.post("/colors")
async def save_colors(colors: Dict[str, ColorConfig]):
    """Save custom color configuration"""
    color_dict = {k: v.model_dump() for k, v in colors.items()}
    await db.color_config.update_one(
        {"type": "shift_colors"},
        {"$set": {"colors": color_dict}},
        upsert=True
    )
    return {"success": True, "colors": color_dict}


@api_router.get("/")
async def root():
    return {"message": "StaffFlow API - Hotel Roster Generator"}


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
